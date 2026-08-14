import io
import sqlite3
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
DB_PATH = "database.db"


def get_connection():
    """إنشاء الاتصال بقاعدة البيانات مع تفعيل المفاتيح الأجنبية"""
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON;")
    return conn


def init_db():
    """تهيئة وإنشاء جميع جداول قاعدة البيانات عند بداية التشغيل"""
    conn = get_connection()
    cursor = conn.cursor()

    # 1. جدول التصنيفات
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL
        )
    """)

    # 2. جدول الصور
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS project_images (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            category_id INTEGER NOT NULL,
            title TEXT NOT NULL,
            image_path TEXT NOT NULL,
            FOREIGN KEY (category_id) REFERENCES categories (id) ON DELETE CASCADE
        )
    """)

    # 3. جدول طلبات المشاريع (تم توحيد اسم العمود إلى name لتفادي التعارض)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS project_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            phone TEXT,
            project_type TEXT,
            details TEXT,
            file_path TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # 4. جدول التقييمات
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS ratings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            client_name TEXT NOT NULL,
            stars INTEGER NOT NULL,
            comment TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # إدراج التصنيفات الافتراضية
    default_cats = [
        "تحويل أكسل",
        "تحويل أكسس",
        "مواقع إلكترونية",
        "أنظمة أخرى",
    ]
    for cat in default_cats:
        cursor.execute(
            "INSERT OR IGNORE INTO categories (name) VALUES (?)", (cat,)
        )

    conn.commit()
    conn.close()

    # تهيئة جدول الإحصائيات وتحديث الهيكل للأعمدة الجديدة
    init_analytics_db()
    update_db_schema()


# ------------------- 📋 دوال طلبات العملاء -------------------


def add_project_request(name, phone, project_type, details, file_path):
    """إضافة طلب جديد وإرجاع رقم الطلب (ID)"""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        """
        INSERT INTO project_requests (name, phone, project_type, details, file_path)
        VALUES (?, ?, ?, ?, ?)
    """,
        (name, phone, project_type, details, file_path),
    )

    request_id = cursor.lastrowid
    conn.commit()
    conn.close()

    return request_id


def get_all_requests():
    """جلب جميع طلبات العملاء"""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT id, name, phone, project_type, details, file_path, created_at FROM project_requests ORDER BY id DESC"
    )
    rows = cursor.fetchall()
    conn.close()
    return rows


def update_db_schema():
    """تحديث جدول الطلبات لإضافة حقول المتابعة والماليات"""
    conn = get_connection()
    cursor = conn.cursor()

    columns = [
        ("status", "TEXT DEFAULT 'طلب جديد'"),
        ("total_amount", "REAL DEFAULT 0.0"),
        ("paid_amount", "REAL DEFAULT 0.0"),
        ("lead_quality", "TEXT DEFAULT 'غير محدد'"),
        ("admin_notes", "TEXT DEFAULT ''"),
        ("is_blacklisted", "INTEGER DEFAULT 0"),
    ]

    for col_name, col_type in columns:
        try:
            cursor.execute(
                f"ALTER TABLE project_requests ADD COLUMN {col_name} {col_type}"
            )
        except sqlite3.OperationalError:
            pass  # العمود موجود مسبقاً

    conn.commit()
    conn.close()


def update_request_details(
    req_id, status, total, paid, quality, notes, is_blacklisted
):
    """تحديث بيانات التتبع المالي والجدية للعميل"""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        """
        UPDATE project_requests 
        SET status=?, total_amount=?, paid_amount=?, lead_quality=?, admin_notes=?, is_blacklisted=?
        WHERE id=?
    """,
        (
            status,
            total,
            paid,
            quality,
            notes,
            1 if is_blacklisted else 0,
            req_id,
        ),
    )
    conn.commit()
    conn.close()


def export_requests_to_excel():
    """تصدير الطلبات إلى ملف Excel بتنسيق احترافي جافاسكريبت/أوبن بايثون"""
    conn = get_connection()
    cursor = conn.cursor()

    # فحص اسم العمود المتاح لضمان عدم حدوث خطأ
    cursor.execute("PRAGMA table_info(project_requests)")
    cols = [row[1] for row in cursor.fetchall()]
    name_col = "name" if "name" in cols else "client_name"

    query = f"""
        SELECT 
            id AS 'المعرف (ID)',
            created_at AS 'تاريخ الطلب',
            {name_col} AS 'الاسم الكامل',
            phone AS 'رقم الهاتف',
            project_type AS 'نوع البيئة',
            status AS 'حالة المشروع',
            lead_quality AS 'جدية العميل',
            total_amount AS 'المبلغ الكلي',
            paid_amount AS 'المبلغ المدفوع',
            (total_amount - paid_amount) AS 'المبلغ المتبقي',
            CASE WHEN is_blacklisted = 1 THEN 'نعم' ELSE 'لا' END AS 'قائمة سوداء',
            details AS 'تفاصيل الطلب',
            admin_notes AS 'ملاحظات الإدارة'
        FROM project_requests
        ORDER BY id DESC
    """
    df = pd.read_sql_query(query, conn)
    conn.close()

    output = io.BytesIO()

    # استخدام openpyxl لتسجيل ورقة العمل وتنسيقها
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="طلبات العملاء")

        # الوصول إلى الكائن المباشر لورقة أكسل
        worksheet = writer.sheets["طلبات العملاء"]

        # 1. ضبط اتجاه الورقة من اليمين إلى اليسار (Right To Left)
        worksheet.sheet_view.rightToLeft = True

        # 2. ألوان وتنسيقات الهيدر والخلايا
        header_fill = PatternFill(
            start_color="1F4E79", end_color="1F4E79", fill_type="solid"
        )  # كحلي احترافي
        header_font = Font(
            name="Calibri", size=12, bold=True, color="FFFFFF"
        )  # خط أبيض عريض
        data_font = Font(name="Calibri", size=11, color="000000")

        center_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        thin_border = Side(style="thin", color="D9D9D9")
        border = Border(
            left=thin_border,
            right=thin_border,
            top=thin_border,
            bottom=thin_border,
        )

        # 3. تطبيق التنسيق على الصف الأول (الهيدر)
        worksheet.row_dimensions[1].height = 28
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

        # 4. تطبيق التنسيق والحدود على بقية البيانات + زيادة ارتفاع الصفوف
        for row in worksheet.iter_rows(
            min_row=2,
            max_row=worksheet.max_row,
            min_col=1,
            max_col=worksheet.max_column,
        ):
            worksheet.row_dimensions[row[0].row].height = 22
            for cell in row:
                cell.font = data_font
                cell.alignment = center_alignment
                cell.border = border

        # 5. ضبط عرض الأعمدة تلقائياً ليناسب محتواها + مساحة إضافية
        for col in worksheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                # معالجة أطوال النصوص العربية
                max_len = max(max_len, len(val_str))

            # إعطاء مساحة مريحة للعمود (بحد أدنى 15 وبحد أقصى 40 للتفاصيل)
            adjusted_width = min(max(max_len + 4, 15), 45)
            worksheet.column_dimensions[col_letter].width = adjusted_width

    return output.getvalue()


# ------------------- 🖼️ دوال التصنيفات والمعرض -------------------


def get_categories():
    """جلب كافة التصنيفات"""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, name FROM categories")
    cats = cursor.fetchall()
    conn.close()
    return cats


def get_unique_titles():
    """جلب كافة النصوص/أوصاف الصور بدون تكرار للفلترة"""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT DISTINCT title FROM project_images WHERE title IS NOT NULL AND title != '' ORDER BY title ASC"
    )
    rows = cursor.fetchall()
    conn.close()
    return [r[0] for r in rows]


def get_images_count_by_category(category_id):
    """جلب عدد الصور لتصنيف محدد"""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT COUNT(*) FROM project_images WHERE category_id = ?",
        (category_id,),
    )
    count = cursor.fetchone()[0]
    conn.close()
    return count


def get_images_count_by_title(title):
    """جلب عدد الصور المضافة لنفس البيان/وصف الصورة (للتحقق من شرط الـ 3 صور)"""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT COUNT(*) FROM project_images WHERE title = ?", (title,)
    )
    count = cursor.fetchone()[0]
    conn.close()
    return count


def add_image_to_category(category_id, title, image_path):
    """إدراج صورة جديدة في قاعدة البيانات"""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO project_images (category_id, title, image_path) VALUES (?, ?, ?)",
        (category_id, title, image_path),
    )
    conn.commit()
    conn.close()


def get_images(filter_value="الكل", filter_by="category"):
    """جلب الصور حسب الفلترة"""
    conn = get_connection()
    cursor = conn.cursor()

    if filter_value == "الكل":
        cursor.execute("""
            SELECT pi.id, pi.title, pi.image_path, c.name 
            FROM project_images pi 
            JOIN categories c ON pi.category_id = c.id
        """)
    elif filter_by == "title":
        cursor.execute(
            """
            SELECT pi.id, pi.title, pi.image_path, c.name 
            FROM project_images pi 
            JOIN categories c ON pi.category_id = c.id 
            WHERE pi.title = ?
        """,
            (filter_value,),
        )
    else:
        cursor.execute(
            """
            SELECT pi.id, pi.title, pi.image_path, c.name 
            FROM project_images pi 
            JOIN categories c ON pi.category_id = c.id 
            WHERE c.name = ?
        """,
            (filter_value,),
        )

    images = cursor.fetchall()
    conn.close()
    return images


def delete_image(image_id):
    """حذف صورة واحدة"""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM project_images WHERE id = ?", (image_id,))
    conn.commit()
    conn.close()


# ------------------- ⭐ دوال نظام التقييمات -------------------


def add_rating(name, stars, comment):
    """إضافة تقييم جديد من عميل"""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO ratings (client_name, stars, comment) VALUES (?, ?, ?)",
        (name, stars, comment),
    )
    conn.commit()
    conn.close()


def get_all_ratings():
    """جلب جميع التقييمات المضافة"""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT id, client_name, stars, comment, created_at FROM ratings ORDER BY id DESC"
    )
    rows = cursor.fetchall()
    conn.close()
    return rows


def get_average_rating():
    """حساب متوسط التقييمات وعدد المقيمين"""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT AVG(stars), COUNT(id) FROM ratings")
    row = cursor.fetchone()
    conn.close()

    avg_stars = round(row[0], 1) if row[0] is not None else 0.0
    total_count = row[1] if row[1] is not None else 0
    return avg_stars, total_count


def delete_rating(rating_id):
    """حذف تقييم بواسطة الإدارة"""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM ratings WHERE id = ?", (rating_id,))
    conn.commit()
    conn.close()


# ------------------- 📊 دوال الإحصائيات والتحليلات -------------------


def init_analytics_db():
    """إنشاء جدول الإحصائيات وتعيين القيم الابتدائية"""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS analytics (
            metric_key TEXT PRIMARY KEY,
            metric_value INTEGER DEFAULT 0
        )
    """)
    cursor.execute(
        "INSERT OR IGNORE INTO analytics (metric_key, metric_value) VALUES ('total_views', 0)"
    )
    cursor.execute(
        "INSERT OR IGNORE INTO analytics (metric_key, metric_value) VALUES ('portfolio_clicks', 0)"
    )
    conn.commit()
    conn.close()


def increment_metric(key_name):
    """زيادة عداد معين بمقدار 1"""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "UPDATE analytics SET metric_value = metric_value + 1 WHERE metric_key = ?",
        (key_name,),
    )
    conn.commit()
    conn.close()


def get_analytics():
    """جلب كافة الإحصائيات مع إجمالي الطلبات"""
    conn = get_connection()
    cursor = conn.cursor()

    init_analytics_db()

    cursor.execute("SELECT metric_key, metric_value FROM analytics")
    data = dict(cursor.fetchall())

    cursor.execute("SELECT COUNT(*) FROM project_requests")
    data["total_requests"] = cursor.fetchone()[0]

    conn.close()
    return data